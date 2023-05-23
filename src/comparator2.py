import numpy as np
import pandas as pd
import math
import openpyxl
import os
from openpyxl.styles import PatternFill

class ComparatorPowerBI:
    def __init__(self, data_path):
        self.data_path = data_path
    
    def unique(self, list):
        res = []
        for x in list:
            if x not in res:
                res.append(x)
        return res
    
    def rearrange_tables(self, df1, df2):
        columns1 = df1.columns.to_list()
        columns2 = df2.columns.to_list()

        unique_columns = self.unique(columns1 + columns2)
        dict1 = {}
        dict2 = {}
        visited = []
        for x in unique_columns:
            if x in df1.columns.to_list() and df2.columns.to_list():
                dict1[x] = list(df1[x].values)
                dict2[x] = list(df2[x].values)
                visited.append(x)

        for x in unique_columns:
            if x in visited:
                continue
            if x in df1.columns.to_list() :
                dict1[x] = list(df1[x].values)
            else:    
                dict2[x] = list(df2[x].values)

        df1 = pd.DataFrame(dict1, columns=dict1.keys())
        df2 = pd.DataFrame(dict2, columns=dict2.keys())

        vis1 = [0 for i in range(len(df1))]
        vis2 = [0 for i in range(len(df2))]

        cnt1 = 0
        cnt2 = 0

        rows1 = []
        rows2 = []
        print(df1)
        print(df2)
        while cnt1 < len(df1) and cnt2 < len(df2):
            best_i = -1
            best_j = -1
            max_match = -1
            for i in range(len(vis1)):    
                if vis1[i] == 0:
                    for j in range(len(vis2)):
                        if vis2[j] == 0:
                            cnt = 0
                            for x in visited:
                                if df1[x][i] == df2[x][j]:
                                    cnt += 1
                            if cnt > max_match:
                                max_match = cnt
                                best_i = i
                                best_j = j
            cnt1 += 1
            cnt2 += 1
            vis1[best_i] = 1
            vis2[best_j] = 1
            rows1.append(df1.iloc[best_i].values.flatten().tolist())
            rows2.append(df2.iloc[best_j].values.flatten().tolist())
        
        for i in range(len(df1)):
            if vis1[i] == 0:
                rows1.append(df1.iloc[i].values.flatten().tolist())

        for i in range(len(df2)):
            if vis2[i] == 0:
                rows2.append(df2.iloc[i].values.flatten().tolist())

        df1 = pd.DataFrame(rows1, columns=df1.columns)
        df2 = pd.DataFrame(rows2, columns=df2.columns)

        return (df1, df2, visited)

    def num_of_digits(self, st):
        cnt = 0
        for x in st:
            if x.isdigit():
                cnt += 1
        return cnt

    def compare_tables(self, df1, df2, common_columns):
        dict_delta = {}
        dict_color = {}
        n_rows1 = len(df1)
        n_rows2 = len(df2)
        n_cols1 = len(df1.columns)
        n_cols2 = len(df2.columns)
        diff_col = 1
        columns = []
        if n_cols1 > n_cols2:
            columns = df1.columns.to_list()
        else:
            columns = df2.columns.to_list()
        for x in columns:
            if x in common_columns:
                dict_delta[x] = []
                dict_color[x] = []
                for i in range(max(len(df1), len(df2))):
                    if i < min(len(df1), len(df2)):
                        value1 = df1[x][i]
                        value2 = df2[x][i]
                        if isinstance(value1, str) and isinstance(value2, str):
                            if len(value1) - self.num_of_digits(value1) > 2 and len(value2) - self.num_of_digits(value2) > 2: 
                                if value1 == value2:
                                    dict_color[x].append('00FF00')
                                    dict_delta[x].append("Both values are strings and are equal")
                                else:
                                    dict_color[x].append('FF0000')
                                    dict_delta[x].append("Both values are strings and are not equal")
                                continue
                            elif len(value1) - self.num_of_digits(value1) > 2 or len(value2) - self.num_of_digits(value2) > 2:
                                dict_color[x].append('FF0000')
                                dict_delta[x].append("One is a number and the other one is a string")
                                continue

                        elif isinstance(value1, str):
                            if len(value1) - self.num_of_digits(value1) > 2:
                                dict_color[x].append('FF0000')
                                dict_delta[x].append("One is a number and the other one is a string")
                                continue
                        elif isinstance(value1, str):
                            if len(value2) - self.num_of_digits(value2) > 2:
                                dict_color[x].append('FF0000')
                                dict_delta[x].append("One is a number and the other one is a string")
                                continue 



                        if not isinstance(value1, str) and not isinstance(value2, str):
                            if math.isnan(value1) or math.isnan(value2):
                                if math.isnan(value1) and math.isnan(value2):
                                    dict_color[x].append('00FF00')
                                    dict_delta[x].append("Both values NaN")
                                else:
                                    dict_color[x].append('FF0000')
                                    dict_delta[x].append("One of the value NaN")
                                continue
                        is_currency1 = ""
                        is_currency2 = ""
                        if isinstance(value1, str):
                            if not value1[0].isdigit():
                                is_currency1 = value1[0]
                                value1 = float(value1[1:])
                            elif not value1[-1].isdigit():
                                is_currency1 = value1[-1]
                                value1 = float(value1[:-2])

                        if isinstance(value2, str):
                            if not value2[0].isdigit():
                                is_currency2 = value2[0]
                                value2 = float(value2[1:])
                            elif not value2[-1].isdigit():
                                is_currency2 = value2[-1]
                                value2 = float(value2[:-2])  
                        value2 = float(value2)
                        value1 = float(value1)
                        diff = abs(value1 - value2)

                        if is_currency1 != is_currency2:
                            dict_color[x].append('FF0000')
                            dict_delta[x].append(str(diff) + " " + "(different currencies)")
                        else:
                            if diff == 0:
                                dict_color[x].append('00FF00')
                            else:
                                dict_color[x].append('FF0000')
                            dict_delta[x].append(str(diff))
                    else:
                        dict_delta[x].append("Different number of rows!")
                        dict_color[x].append("FF0000")
            else:
                dict_color["diff_column" + "_" + str(diff_col)] = []
                dict_delta["diff_column" + "_" + str(diff_col)] = []
                for i in range(max(len(df1), len(df2))):
                    dict_color["diff_column" + "_" + str(diff_col)].append('FF0000')
                    dict_delta["diff_column" + "_" + str(diff_col)].append("Different columns")
                diff_col += 1 

        delta_df = pd.DataFrame(dict_delta, columns=dict_delta.keys())
        delta_df_np = pd.DataFrame(dict_color, columns=dict_color.keys()).to_numpy()
        return df1, df2, n_rows1, n_rows2, n_cols1, n_cols2, delta_df, delta_df_np



    def compare_pages(self, path1, path2):
        l_df1 = os.listdir(path1)
        l_df2 = os.listdir(path2)

        l_df1 = sorted(l_df1)
        l_df2 = sorted(l_df2)

        list_df1 = []
        list_df2 = []
        list_n_rows1 = []
        list_n_rows2 = []
        list_n_cols1 = []
        list_n_cols2 = []
        list_deltas = []
        list_deltas_np = []
        names1 = []
        names2 = []

        for (df1_name, df2_name) in zip(l_df1, l_df2):
            df1 = pd.read_csv(f"{path1}/{df1_name}")
            df2 = pd.read_csv(f"{path2}/{df2_name}")
            names1.append(df1_name[:-4])
            names2.append(df2_name[:-4])
            (df1, df2, common_columns) = self.rearrange_tables(df1, df2)
            df1, df2, n_rows1, n_rows2, n_cols1, n_cols2, deltas, deltas_np = self.compare_tables(df1, df2, common_columns)
            list_df1.append(df1)
            list_df2.append(df2)
            list_n_rows1.append(n_rows1)
            list_n_rows2.append(n_rows2)
            list_n_cols1.append(n_cols1)
            list_n_cols2.append(n_cols2)
            list_deltas.append(deltas)
            list_deltas_np.append(deltas_np)
        return list_df1, list_df2, list_n_rows1, list_n_rows2, list_n_cols1, list_n_cols2, list_deltas, list_deltas_np, names1, names2


    def create_predictions(self):
        viz1 = os.listdir(f"{self.data_path}/Report1")
        viz2 = os.listdir(f"{self.data_path}/Report2")

        viz1 = sorted(viz1)
        viz2 = sorted(viz2)
        workbook = openpyxl.Workbook()
        for (page1, page2) in zip(viz1, viz2):
            if len(os.listdir(f"{self.data_path}/Report1/{page1}")) != len(os.listdir(f"{self.data_path}/Report2/{page2}")):
                print("THESE ARE DIFFERENT REPORTS")
                return
            sheet = workbook.create_sheet(title=".".join(page1.split(".")[1:]))
            list_df1, list_df2, list_n_rows1, list_n_rows2, list_n_cols1, list_n_cols2, list_deltas, list_deltas_np, names1, names2 = self.compare_pages(f"{self.data_path}/Report1/{page1}", f"{self.data_path}/Report2/{page2}")
            
            start_x = "A"
            start_y = 4
            inc_x = 0
            inc_y = 0
            

            sheet["A1"] = f"Comparison of page {sheet.title}"
            sheet["A2"] = "Report1"
            sheet[f"{chr(ord('A') + inc_x)}2"] = "Report2"
            
            for (df1, df2, nr1, nr2, nc1, nc2, d, d_np, n1, n2) in zip(list_df1, list_df2, list_n_rows1, list_n_rows2, list_n_cols1, list_n_cols2, list_deltas, list_deltas_np, names1, names2):
                sheet["A3"] = ".".join(n1.split(".")[1:])
                #print("cole:", chr(ord('A') + int(len(df1)) + 3))
                sheet[f"{chr(ord('A') + int(len(df1.columns)) + 3)}3"] = ".".join(n2.split(".")[1:])

                start_row = start_y
                start_col = start_x
                inc_y = max(len(df1), len(df2)) + 3
                inc_x = len(df1.columns) + 3
                col_ = start_col
                row_ = start_row
                for val in df1.columns:
                    sheet[f'{col_}{row_}'] = val
                    col_ = chr(ord(col_) + 1)
                row_ += 1
                for idx, row in df1.iterrows():
                    col_ = start_col
                    for cell in row:
                        sheet[f'{col_}{row_}'] = cell
                        col_ = chr(ord(col_) + 1)
                    row_ += 1

                sheet[f"{start_col}{start_y + inc_y}"] = "Number of rows"
                sheet[f"{start_col}{start_y + inc_y + 2}"] = "Number of columns"
                sheet[f"{start_col}{start_y + inc_y + 1}"] = nr1
                sheet[f"{start_col}{start_y + inc_y + 3}"] = nc1

                start_col = chr(ord(start_col) + inc_x)
                start_row = start_y
                inc_x = len(df2.columns) + 3

                col_ = start_col
                row_ = start_row
                print("Pocetni:", col_, row_)
                for val in df2.columns:
                    print(col_, row_)
                    sheet[f'{col_}{row_}'] = val
                    col_ = chr(ord(col_) + 1)
                row_ += 1
                for idx, row in df2.iterrows():
                    col_ = start_col
                    for cell in row:
                        sheet[f'{col_}{row_}'] = cell
                        col_ = chr(ord(col_) + 1)
                    row_ += 1

                sheet[f"{start_col}{start_y + inc_y}"] = "Number of rows"
                sheet[f"{start_col}{start_y + inc_y + 2}"] = "Number of columns"
                sheet[f"{start_col}{start_y + inc_y + 1}"] = nr2
                sheet[f"{start_col}{start_y + inc_y + 3}"] = nc2

                start_col = chr(ord(start_col) + inc_x)
                start_row = start_y
                inc_x = len(df1.columns) + 3
                sheet[f"{start_col}{start_row - 1}"] = "Differences between Vizuals"

                col_ = start_col
                row_ = start_row
                is_time_col = False
                i = 0
                j = 0
                for val in d.columns:
                    sheet[f'{col_}{row_}'] = "Delta " + val
                    col_ = chr(ord(col_) + 1)
                row_ += 1
                for idx, row in d.iterrows():
                    j = 0
                    col_ = start_col
                    for cell in row:
                        sheet[f'{col_}{row_}'] = cell
                        sheet[f'{col_}{row_}'].fill = PatternFill(start_color=d_np[i, j], end_color=d_np[i, j], fill_type='solid')
                        j += 1
                        col_ = chr(ord(col_) + 1)
                    row_ += 1
                    i += 1


                sheet[f"{start_col}{start_y + inc_y}"] = "Delta Number of rows"
                sheet[f"{start_col}{start_y + inc_y + 2}"] = "Delta Number of columns"
                sheet[f"{start_col}{start_y + inc_y + 1}"] = abs(nr1 - nr2)
                sheet[f"{start_col}{start_y + inc_y + 3}"] = abs(nc1 - nc2)
                col1 = "FF0000"
                if nr1 == nr2:
                    col1 = "00FF00"
                col2 = "FF0000"
                if nc1 == nc2:
                    col2 = "00FF00"
                sheet[f"{start_col}{start_y + inc_y + 1}"].fill = PatternFill(start_color=col1, end_color=col1, fill_type='solid')
                sheet[f"{start_col}{start_y + inc_y + 3}"].fill = PatternFill(start_color=col2, end_color=col2, fill_type='solid')

                start_y += inc_y + 7

        workbook.save(f'./{self.data_path}/comparisonReport.xlsx')

        return "SUCCESSFULY CREATED REPORT"







