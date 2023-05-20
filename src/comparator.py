import numpy as np
import pandas as pd
import math
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
class PowerBIComparator:
    def __init__(self, data_path1, data_path2):
        self.path1 = data_path1
        self.path2 = data_path2
    def unique(self, list):
        res = []
        for x in list:
            if x not in res:
                res.append(x)

        return res

    def compare(self, df1, df2, key_word, key_id, new_feature=0):
        n_rows_1 = len(df1)
        n_rows_2 = len(df2)
        n_cols_1 = len(df1.columns)
        n_cols_2 = len(df2.columns)
        unique_columns = self.unique(df1.columns.to_list() + df2.columns.to_list())
        print(unique_columns)
        dict1 = {}
        dict2 = {}
        visited = []
        for x in unique_columns:
            if x in df1.columns.to_list() and df2.columns.to_list():
                dict1[x] = list(df1[x].values)
                dict2[x] = list(df2[x].values)
                visited.append(x)

        for x in unique_columns:
            if x in visited:
                continue
            if x in df1.columns.to_list() :
                dict1[x] = list(df1[x].values)
            else:    
                dict2[x] = list(df2[x].values)

        df1 = pd.DataFrame(dict1, columns=dict1.keys())
        df2 = pd.DataFrame(dict2, columns=dict2.keys())

        unique_vals = self.unique(df1[key_word].to_list() + df2[key_word].to_list())

        rows1 = []
        rows2 = []
        n_equal = 0
        visited = []
        for x in unique_vals:
            if x in df1[key_word].to_list() and x in df2[key_word].to_list():
                visited.append(x)
                rows1.append(df1.loc[df1[key_word] == x].values.flatten().tolist())
                rows2.append(df2.loc[df2[key_word] == x].values.flatten().tolist())
        
        for x in unique_vals:
            if x in visited:
                continue
            if x in df1[key_word].to_list():
                rows1.append(df1.loc[df1[key_word] == x].values.flatten().tolist())
            else:
                rows2.append(df2.loc[df2[key_word] == x].values.flatten().tolist())

        df1_new = pd.DataFrame(rows1, columns=df1.columns)
        df2_new = pd.DataFrame(rows2, columns=df2.columns)

        if new_feature == 1:
            df1_new = df1_new.drop(df1_new.columns[-1], axis=1)
            df2_new = df2_new.drop(df2_new.columns[-1], axis=1)

        key_id_new = []
        for x in key_id:
            df1_new = df1_new.rename(columns={
                x : f"KEY_ID[{x}]"
            })
            df2_new = df2_new.rename(columns={
                x : f"KEY_ID[{x}]"
            })
            key_id_new.append(f"KEY_ID[{x}]")
        dict_delta = {}
        dict_color = {}
        diff_col = 1
        for x in self.unique(df1_new.columns.to_list() + df2_new.columns.to_list()):
            if x in df1_new.columns.to_list() and x in df2_new.columns.to_list():
                dict_color[x] = []
                dict_delta[x] = []
                for i in range(max(len(df1_new), len(df2_new))):
                   

                    if i < min(len(df1_new), len(df2_new)):
                        if x in key_id_new:
                            value1 = df1_new[x][i]
                            value2 = df2_new[x][i]

                            if value1 == value2:
                                dict_color[x].append('00FF00')
                                dict_delta[x].append("Same key values")
                            else:
                                dict_color[x].append('00FF00')
                                dict_delta[x].append("Different key values")
                            continue

                        value1 = df1_new[x][i]
                        value2 = df2_new[x][i]
                        if not isinstance(value1, str) and not isinstance(value2, str):
                            if math.isnan(value1) or math.isnan(value2):
                                if math.isnan(value1) and math.isnan(value2):
                                    dict_color[x].append('00FF00')
                                    dict_delta[x].append("Both values NaN")
                                else:
                                    dict_color[x].append('FF0000')
                                    dict_delta[x].append("One of the value NaN")
                                continue
                        is_currency1 = ""
                        is_currency2 = ""
                        if isinstance(value1, str):
                            if not value1[0].isdigit():
                                is_currency1 = value1[0]
                                value1 = float(value1[1:])
                            elif not value1[-1].isdigit():
                                is_currency1 = value1[-1]
                                value1 = float(value1[:-2])

                        if isinstance(value2, str):
                            if not value2[0].isdigit():
                                is_currency2 = value2[0]
                                value2 = float(value2[1:])
                            elif not value2[-1].isdigit():
                                is_currency2 = value2[-1]
                                value2 = float(value2[:-2])  
                        value2 = float(value2)
                        value1 = float(value1)
                        diff = abs(value1 - value2)

                        if is_currency1 != is_currency2:
                            dict_color[x].append('FF0000')
                            dict_delta[x].append(str(diff) + " " + "(different currencies)")
                        else:
                            if diff == 0:
                                dict_color[x].append('00FF00')
                            else:
                                dict_color[x].append('FF0000')
                            dict_delta[x].append(str(diff))
                    else:
                        dict_color[x].append('FF0000')
                        dict_delta[x].append("Different key values")

            else:
                dict_color["diff_column" + "_" + str(diff_col)] = []
                dict_delta["diff_column" + "_" + str(diff_col)] = []
                for i in range(max(len(df1_new), len(df2_new))):
                    dict_color["diff_column" + "_" + str(diff_col)].append('FF0000')
                    dict_delta["diff_column" + "_" + str(diff_col)].append("Different columns")
                dif_col += 1 

        dict_num = {}
        dict_num_col = {}
        
        dict_num["Number of rows"] = [abs(n_rows_1 - n_rows_2)]
        if n_rows_2 != n_rows_1:
            dict_num_col["Number of rows"] = ['FF0000']
        else:
            dict_num_col["Number of rows"] = ['00FF00']

        dict_num["Number of columns"] = [abs(n_cols_1 - n_cols_2)]
        if n_cols_2 != n_cols_1:
            dict_num_col["Number of columns"] = ['FF0000']
        else:
            dict_num_col["Number of columns"] = ['00FF00']    

        delta_df = pd.DataFrame(dict_delta, columns=dict_delta.keys())
        delta_df_np = pd.DataFrame(dict_color, columns=dict_color.keys()).to_numpy()
        num_df = pd.DataFrame(dict_num, columns=dict_num.keys())
        num_df_np = pd.DataFrame(dict_num_col, columns=dict_num_col.keys())
        return df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np
    
    
    def convert_to_normal_csv():
        pass

    def input_tables(self):

        sheets = {}

        df1 = pd.read_csv(self.path1 + "/Sales quantity.csv")
        df2 = pd.read_csv(self.path2 + "/Sales quantity.csv")
        new_list1 = []
        for i in range(len(df1)):
            new_list1.append(str(df1["Month short"][i] + str(df1["Date week"][i])))
        df1["Week"] = new_list1
        new_list2 = []
        for i in range(len(df2)):
            new_list2.append(str(df2["Month short"][i] + str(df2["Date week"][i])))
        df2["Week"] = new_list2
        df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np = self.compare(df1, df2, "Week", key_id=["Month short", "Date week"], new_feature=1)
        sheets["Sales quantity"] = [df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np]
        # Sales quantity veggie
        df1 = pd.read_csv(self.path1 + "/Sales quantity by veggie.csv")
        df2 = pd.read_csv(self.path2 + "/Sales quantity by veggie.csv")

        df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np = self.compare(df1, df2, "Veggie", key_id=["Veggie"], new_feature=0)

        sheets["Sales quantity by veggie"] = [df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np]
        # Sales quantity customer
        df1 = pd.read_csv(self.path1 + "/Sales quantity by customer.csv")
        df2 = pd.read_csv(self.path2 + "/Sales quantity by customer.csv")

        df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np = self.compare(df1, df2, "Customer", key_id=["Customer"], new_feature=0)
        sheets["Sales quantity by customer"] = [df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np]
        # Sales quantity country
        df1 = pd.read_csv(self.path1 + "/Sales quantity by country.csv")
        df2 = pd.read_csv(self.path2 + "/Sales quantity by country.csv")

        df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np = self.compare(df1, df2, "Country", key_id=["Country"], new_feature=0)
        sheets["Sales quantity by country"] = [df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np]
        # Sales quantity SY VS PY
        df1 = pd.read_csv(self.path1 + "/SY VS PY.csv")
        df2 = pd.read_csv(self.path2 + "/SY VS PY.csv")

        df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np = self.compare(df1, df2, "Date year", key_id=["Date year"], new_feature=0)
        sheets["SY vs PY Quantity"] = [df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np]
        df1 = pd.read_csv(self.path1 + "/Sales value.csv")
        df2 = pd.read_csv(self.path2 + "/Sales value.csv")
        new_list1 = []
        for i in range(len(df1)):
            new_list1.append(str(df1["Month short"][i] + str(df1["Date week"][i])))
        df1["Week"] = new_list1
        new_list2 = []
        for i in range(len(df2)):
            new_list2.append(str(df2["Month short"][i] + str(df2["Date week"][i])))
        df2["Week"] = new_list2
        df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np = self.compare(df1, df2, "Week", key_id=["Month short", "Date week"], new_feature=1)
        sheets["Sales value"] = [df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np]
        # Sales value veggie
        df1 = pd.read_csv(self.path1 + "/Sales value by veggie.csv")
        df2 = pd.read_csv(self.path2 + "/Sales value by veggie.csv")

        df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np = self.compare(df1, df2, "Veggie", key_id=["Veggie"], new_feature=0)
        sheets["Sales value by veggie"] = [df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np]
        # Sales value customer
        df1 = pd.read_csv(self.path1 + "/Sales value by customer.csv")
        df2 = pd.read_csv(self.path2 + "/Sales value by customer.csv")

        df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np = self.compare(df1, df2, "Customer", key_id=["Customer"], new_feature=0)
        sheets["Sales value by customer"] = [df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np]
        # Sales value country
        df1 = pd.read_csv(self.path1 + "/Sales value by country.csv")
        df2 = pd.read_csv(self.path2 + "/Sales value by country.csv")

        df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np = self.compare(df1, df2, "Country", key_id=["Country"], new_feature=0)
        sheets["Sales value by country"] = [df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np]
        # Sales value SY VS PY
        df1 = pd.read_csv(self.path1 + "/SY VS PY (1).csv")
        df2 = pd.read_csv(self.path2 + "/SY VS PY (1).csv")

        df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np = self.compare(df1, df2, "Date year", key_id=["Date year"], new_feature=0)
        sheets["SY vs PY Value"] = [df1_new, df2_new, n_rows_1, n_rows_2, n_cols_1, n_cols_2, delta_df, delta_df_np, num_df, num_df_np]

        dict_status = {}
        dict_status_col = {}
        dict_status["Visual"] = []
        dict_status["Status"] = []
        dict_status_col["Status"] = []
        for key in sheets.keys():
            lst = np.unique(np.concatenate([np.unique(sheets[key][7]), np.unique(sheets[key][9])]))
            if len(lst) == 1 and lst[0] == "00FF00":
                dict_status["Visual"].append(key)
                dict_status["Status"].append("Match")
                dict_status_col["Status"].append("00FF00")
            else:
                dict_status["Visual"].append(key)
                dict_status["Status"].append("Not Match")
                dict_status_col["Status"].append("FF0000")
        sheets["Overview"] = [pd.DataFrame(dict_status, columns=dict_status.keys()), pd.DataFrame(dict_status_col, columns=dict_status_col.keys()).to_numpy()]
        self.create_results(sheets)
    def create_results(self, sheets):

        workbook = openpyxl.Workbook()
        for i, x in enumerate(["Overview", "SY vs PY Quantity", "Sales quantity by country", "Sales quantity by customer", "Sales quantity by veggie", "Sales quantity", "SY vs PY Value", "Sales value by country", "Sales value by customer", "Sales value by veggie", "Sales value"]):
            sheet = workbook.create_sheet(title=x)

            if x == "Overview":
                start_row = 2  
                start_column = 'A'  
                sheet["A1"] = "Overview"

                col_ = start_column
                row_ = start_row
                is_time_col = False
                i = 0
                j = 0
                for val in sheets[x][0].columns:
                    sheet[f'{col_}{row_}'] = val
                    col_ = chr(ord(col_) + 1)
                row_ += 1
                for idx, row in sheets[x][0].iterrows():
                    is_time_col = False
                    j = 0
                    col_ = start_column
                    for cell in row:
                        sheet[f'{col_}{row_}'] = cell
                        
                        if is_time_col == True:
                            sheet[f'{col_}{row_}'].fill = PatternFill(start_color=sheets[x][1][i, j], end_color=sheets[x][1][i, j], fill_type='solid')
                            j += 1
                        col_ = chr(ord(col_) + 1)
                        is_time_col = True
                    row_ += 1
                    i += 1

            else:
                
                sheet["A1"] = "Comparison of " + x
                sheet["A2"] = "Report 1"
                ch = chr(len(sheets[x][0].columns) + 3 + 64)

                sheet[f"{ch}2"] = "Report 2"

                start_row = 3  
                start_column = 'A'  
                col_ = start_column
                row_ = start_row
                for val in sheets[x][0].columns:
                    sheet[f'{col_}{row_}'] = val
                    col_ = chr(ord(col_) + 1)
                row_ += 1
                for idx, row in sheets[x][0].iterrows():
                    col_ = start_column
                    for cell in row:
                        sheet[f'{col_}{row_}'] = cell
                        col_ = chr(ord(col_) + 1)
                    row_ += 1

                sheet[f"{start_column}{row_ + 2}"] = "Number of rows"
                sheet[f"{start_column}{row_ + 4}"] = "Number of columns"
                sheet[f"{start_column}{row_ + 3}"] = sheets[x][2]
                sheet[f"{start_column}{row_ + 5}"] = sheets[x][4]

                start_row = 3  
                start_column = ch
                col_ = start_column
                row_ = start_row
                for val in sheets[x][1].columns:
                    sheet[f'{col_}{row_}'] = val
                    col_ = chr(ord(col_) + 1)
                row_ += 1
                for idx, row in sheets[x][1].iterrows():
                    col_ = start_column
                    for cell in row:
                        sheet[f'{col_}{row_}'] = cell
                        col_ = chr(ord(col_) + 1)
                    row_ += 1

                sheet[f"{start_column}{row_ + 2}"] = "Number of rows"
                sheet[f"{start_column}{row_ + 4}"] = "Number of columns"
                sheet[f"{start_column}{row_ + 3}"] = sheets[x][3]
                sheet[f"{start_column}{row_ + 5}"] = sheets[x][5]

                start_row = 3  
                start_column = chr(ord(ch) + len(sheets[x][1].columns) + 3)
                sheet[f"{start_column}2"] = "Delta between the reports"

                col_ = start_column
                row_ = start_row
                is_time_col = False
                i = 0
                j = 0
                for val in sheets[x][6].columns:
                    sheet[f'{col_}{row_}'] = "Delta " + val
                    col_ = chr(ord(col_) + 1)
                row_ += 1
                for idx, row in sheets[x][6].iterrows():
                    j = 0
                    col_ = start_column
                    for cell in row:
                        sheet[f'{col_}{row_}'] = cell
                        sheet[f'{col_}{row_}'].fill = PatternFill(start_color=sheets[x][7][i, j], end_color=sheets[x][7][i, j], fill_type='solid')
                        j += 1
                        col_ = chr(ord(col_) + 1)
                    row_ += 1
                    i += 1

                sheet[f"{start_column}{row_ + 2}"] = "Delta Number of rows"
                sheet[f"{start_column}{row_ + 4}"] = "Delta Number of columns"
                sheet[f"{start_column}{row_ + 3}"] = sheets[x][8]["Number of rows"][0]
                sheet[f"{start_column}{row_ + 5}"] = sheets[x][8]["Number of columns"][0]
                sheet[f"{start_column}{row_ + 3}"].fill = PatternFill(start_color=sheets[x][9]["Number of rows"][0], end_color=sheets[x][9]["Number of rows"][0], fill_type='solid')
                sheet[f"{start_column}{row_ + 5}"].fill = PatternFill(start_color=sheets[x][9]["Number of columns"][0], end_color=sheets[x][9]["Number of columns"][0], fill_type='solid')

                

        sheet_to_delete = workbook['Sheet']

        # Delete the sheet
        workbook.remove(sheet_to_delete)
        workbook.save('./predictions/comparisonReport.xlsx')

